from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def validate_header(header):
    """
    :param header: list
    :return: header if header is list.
    """
    if isinstance(header, list):
        return header
    else:
        raise ValueError('Header isin´t list type')


def validate_data(data_keys):
    """
    :param data_keys: list
    :return: data_keys if is list.
    """
    if isinstance(data_keys, list):
        return data_keys
    else:
        raise ValueError('Param isin´t list type')


def generate_header(header, ws, start_line):
    """
    :param header: list
    :param ws: file (worksheet)
    :param start_line: start_line: int, represent which line header will start.
    :return:
    """
    header_keys = validate_header(header)

    row_number = start_line
    for col, value in enumerate(header_keys, start=1):
        ws.cell(column=col, row=row_number, value=value)


def gerenerate_data(data, data_keys, ws, start_line):
    """
    :param data: dict
    :param data_keys: list
    :param ws: file (worksheet)
    :param start_line: int, represent which line data will start.
    :return: ws after write rows.
    """
    data_keys = validate_data(data_keys)

    row_number = start_line
    for row in data:
        for col, row_value in enumerate(data_keys, start=1):
            ws.cell(column=col, row=row_number, value=row.get(row_value))
        row_number += 1

    return ws

